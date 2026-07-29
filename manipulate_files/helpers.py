import os
from pathlib import Path
from openpyxl import load_workbook
import shutil
import zipfile

def delete_zip_files(folder):
    for filename in os.listdir(folder):
        if filename.endswith(".zip"):
            os.remove(os.path.join(folder, filename))

def format_date_column(filepath, columns, sheet=None):
    wb = load_workbook(filepath)
    ws = wb[sheet] if sheet else wb.active
    for col in columns:
        for cell in ws[col]:
            cell.number_format = "mm/dd/yyyy"
    wb.save(filepath)

def move_files(source_folder, destination_folder):
    for filename in os.listdir(source_folder):
        shutil.move(os.path.join(source_folder, filename), os.path.join(destination_folder, filename))

def unzip_files(source_folder, destination_folder):
    for filename in os.listdir(source_folder):
        if filename.endswith(".zip"):
            with zipfile.ZipFile(os.path.join(source_folder, filename)) as zf:
                zf.extractall(destination_folder)

