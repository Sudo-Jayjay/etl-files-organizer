from openpyxl import load_workbook

def format_date_column(filepath, columns, sheet=None):
    wb = load_workbook(filepath)
    ws = wb[sheet] if sheet else wb.active
    for col in columns:
        for cell in ws[col]:
            cell.number_format = "mm/dd/yyyy"
    wb.save(filepath)